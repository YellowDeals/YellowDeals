from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.urls import reverse
from django.utils.timezone import now
from django.conf import settings
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.db.models import Sum
from django.db.models import Case, When
from datetime import timedelta
from django.db.models import F
from django.contrib.auth.decorators import login_required
import locale
import xlsxwriter
import re ,random
import pytz , json
from django.contrib import messages 
from .models import Categories, TypeStore,StoreOwner,Product, Admin, GeneralUser, StoreOwner,Complaint
from .models import TypeComplaint,SavedProduct,TypeProduct,SavedProduct
import requests, geopy 
from geopy.distance import geodesic
from django.db.models import Q, Min
from urllib.parse import urlencode
from django.contrib.auth import login, authenticate
from .forms import CustomAuthenticationForm,GeneralUserRegistrationForm ,StoreOwnerRegistrationForm,AdminRegisterForm
from .forms import TypeComplaintForm, TypeProductForm, TypeStoreForm, CategoriesForm,CustomSetPasswordForm,ProductForm,StoreProfileForm
from django.contrib.auth.hashers import make_password
from django.contrib.auth.hashers import check_password
from django.contrib.auth import logout
from django.contrib import messages
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.contrib import messages
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from datetime import datetime, timedelta
import urllib.parse
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.core.mail import send_mail, BadHeaderError
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .utils import calculate_distance
from geopy.distance import geodesic  
from django.http import HttpResponseRedirect



def calculate_distance(user_lat, user_lng, store_lat, store_lng):
    user_location = (user_lat, user_lng)
    store_location = (store_lat, store_lng)
    return geodesic(user_location, store_location).km

def home(request):
    try:
        user_lat = float(request.GET.get('user_lat', 0))
        user_lng = float(request.GET.get('user_lng', 0))
    except ValueError:
        user_lat, user_lng = 0.0, 0.0

    distance_range = request.GET.get('distance_range')
    distance_range = float(distance_range) if distance_range else None

    category = request.GET.get('category', 'all')
    store_type = request.GET.get('store_type', '')

    categories = Categories.objects.all()
    store_types = TypeStore.objects.all()
    product_types = TypeProduct.objects.all()

    username = request.session.get('username')
    is_admin = Admin.objects.filter(admin_username=username).exists()
    is_store_owner = StoreOwner.objects.filter(store_username=username).exists()
    is_general_user = GeneralUser.objects.filter(customer_username=username).exists()

    filters = Q()
    if store_type:
        filters &= Q(typeStore__typeStore_id=store_type)

    if category != 'all':
        store_ids_in_category = Product.objects.filter(
            categories_id=category
        ).values_list('store_owner_id', flat=True).distinct()
        filters &= Q(store_owner_id__in=store_ids_in_category)

    stores = StoreOwner.objects.filter(filters)

    nearby_stores = []
    nearby_store_ids = []

    for store in stores:
        if all([store.latitude, store.longitude]) and user_lat != 0 and user_lng != 0 and distance_range is not None:
            try:
                distance = geodesic((user_lat, user_lng), (store.latitude, store.longitude)).km
                if distance <= distance_range:
                    nearby_stores.append({
                        'name': store.store_name,
                        'address': store.store_address,
                        'distance': round(distance, 2),
                        'category': 'N/A',
                        'store_type': store.typeStore.typeStore_name,
                    })
                    nearby_store_ids.append(store.store_owner_id)
            except Exception as e:
                print(f"Error calculating distance for store {store.store_name}: {e}")
        else:
            nearby_store_ids.append(store.store_owner_id)

    now = timezone.now()
    products_queryset = Product.objects.filter(
        Q(end_time__isnull=True) | Q(end_time__gte=now),
        is_hidden=False
    )

    search_query = request.GET.get('q', '').strip()
    product_type_id = request.GET.get('product_type')
    price_range = request.GET.get('price_range')
    sale_date = request.GET.get('sale_date')

    # ✅ ตัวแปรที่ใช้ใน template
    has_distance = distance_range is not None
    compare_price_enabled = any([
        search_query, category != 'all', product_type_id, store_type,
        price_range, sale_date,
    ])

    has_filter = has_distance or compare_price_enabled

    if has_filter:
        if nearby_store_ids:
            products_queryset = products_queryset.filter(store_owner_id__in=nearby_store_ids)
        else:
            products_queryset = Product.objects.none()

        if search_query:
            keywords = search_query.split()
            query_filter = Q()
            for kw in keywords:
                query_filter |= (
                    Q(product_name__icontains=kw) |
                    Q(store_owner__store_name__icontains=kw) |
                    Q(store_owner__typeStore__typeStore_name__icontains=kw) |
                    Q(product_type__productType_name__icontains=kw) |
                    Q(categories__categories_name__icontains=kw)
                )
            products_queryset = products_queryset.filter(query_filter)

        if category != 'all':
            products_queryset = products_queryset.filter(categories_id=category)

        if product_type_id:
            products_queryset = products_queryset.filter(product_type_id=product_type_id)

        if store_type:
            products_queryset = products_queryset.filter(store_owner__typeStore_id=store_type)

        if price_range and '-' in price_range:
            try:
                min_price_range, max_price_range = map(float, price_range.split('-'))
                products_queryset = products_queryset.filter(
                    product_price__gte=min_price_range,
                    product_price__lte=max_price_range
                )
            except ValueError:
                pass

        if sale_date:
            products_queryset = products_queryset.filter(product_date__date=sale_date)
    else:
        products_queryset = Product.objects.filter(
        Q(end_time__isnull=True) | Q(end_time__gte=now),
        is_hidden=False
    ).order_by('?')[:50]

    product_name_min_prices = (
        Product.objects
        .filter(is_hidden=False)
        .values('product_name')
        .annotate(min_price=Min('product_price'))
    )
    name_to_min_price = {entry['product_name']: entry['min_price'] for entry in product_name_min_prices}

    product_data = []
    for product in products_queryset:
        distance = None
        if user_lat and user_lng and product.store_owner.latitude and product.store_owner.longitude:
            try:
                distance = geodesic(
                    (user_lat, user_lng),
                    (product.store_owner.latitude, product.store_owner.longitude)
                ).km
            except:
                pass

        min_price_for_product = name_to_min_price.get(product.product_name)
        price_diff = None
        if min_price_for_product is not None and product.product_price > min_price_for_product:
            price_diff = round(product.product_price - min_price_for_product, 2)

        product_data.append({
            'product': product,
            'price_diff': price_diff,
            'distance': round(distance, 2) if distance else None,
            'min_price': min_price_for_product,
        })

    paginator = Paginator(product_data, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'home.html', {
        'categories': categories,
        'store_types': store_types,
        'product_types': product_types,
        'stores': nearby_stores,
        'selected_category': category,
        'selected_store_type': store_type,
        'page_obj': page_obj,
        'search_query': search_query,
        'is_admin': is_admin,
        'is_store_owner': is_store_owner,
        'is_general_user': is_general_user,
        'username': username,
        'user_lat': user_lat,
        'user_lng': user_lng,

        # ✅ เพิ่มตัวแปรใหม่เข้า context
        'has_distance': has_distance,
        'compare_price_enabled': compare_price_enabled,
    })
#------- หน้ารายงานการใช้งาน ----------
def submit_complaint_view(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')
    
    general_user = GeneralUser.objects.filter(customer_username=username).first()
    store_user = StoreOwner.objects.filter(store_username=username).first()

    if not general_user and not store_user:
        return redirect('login')

    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        type_id = request.POST.get('type_id')
        text = request.POST.get('complaint_text')
        image = request.FILES.get('complaint_image')

        complaint_type = TypeComplaint.objects.get(pk=type_id)
        target_store = None

        if report_type == 'store':
            store_id = request.POST.get('store_id')
            if store_id:
                target_store = StoreOwner.objects.get(pk=store_id)

        # สร้างคำร้องโดยกำหนดผู้ร้องคนใดคนหนึ่ง
        Complaint.objects.create(
            customer=general_user if general_user else None,
            store_complainer=store_user if store_user else None,
            store_owner=target_store,
            complaint_text=text,
            complaint_status="รอดำเนินการ",
            type_complaint=complaint_type,
            complaint_date=timezone.now(),
            complaint_image=image,
        )

        return render(request, 'submit_complaint.html', {
            'store_owners': StoreOwner.objects.all(),
            'types': TypeComplaint.objects.all(),
            'message': 'ส่งรายงานเรียบร้อยแล้ว'
        })

    return render(request, 'submit_complaint.html', {
        'store_owners': StoreOwner.objects.all(),
        'types': TypeComplaint.objects.all()
    })

#---หน้ารายละเอียดสินค้าที่บันทึก----
def saved_products_view(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')
    
    is_admin = Admin.objects.filter(admin_username=username).exists()
    is_store_owner = StoreOwner.objects.filter(store_username=username).exists()

    user = GeneralUser.objects.filter(customer_username=username).first()
    if not user:
        return redirect('login')

    saved_products = SavedProduct.objects.filter(customer=user)

    category = request.GET.get('category')
    if category:
        saved_products = saved_products.filter(product__product_type__productType_name=category)

    categories = TypeProduct.objects.values_list('productType_name', flat=True).distinct()

    next_url = request.get_full_path()  # เก็บ path ปัจจุบันเป็น next

    return render(request, 'saved_products.html', {
        'is_admin': is_admin,
        'is_store_owner': is_store_owner,
        'saved_products': saved_products,
        'categories': categories,
        'selected_category': category,
        'next_url': next_url,
    })


def remove_saved_product(request, saved_id):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    user = GeneralUser.objects.filter(customer_username=username).first()
    if not user:
        return redirect('login')

    next_url = request.GET.get('next', '/saved_products/')
    if '?' in next_url:
        next_url += '&deleted=1'
    else:
        next_url += '?deleted=1'

    try:
        saved = SavedProduct.objects.get(pk=saved_id, customer=user)
        saved.delete()
    except SavedProduct.DoesNotExist:
        pass  # จะไม่ลบก็ไม่เป็นไร ให้ redirect กลับเฉย ๆ
    return redirect(next_url)

#-------รายละเอียดสินค้าและการบันทึก--------
def product_detail(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)

    viewed_products = request.session.get('viewed_products', [])

    if product_id not in viewed_products:
        Product.objects.filter(product_id=product_id).update(view_count=F('view_count') + 1)
        viewed_products.append(product_id)
        request.session['viewed_products'] = viewed_products

    username = request.session.get('username')
    is_admin = Admin.objects.filter(admin_username=username).exists()
    is_store_owner = StoreOwner.objects.filter(store_username=username).exists()
    user = GeneralUser.objects.filter(customer_username=username).first()

    can_save = user and not (is_admin or is_store_owner)

    saved = False
    if user:
        saved = SavedProduct.objects.filter(customer=user, product=product).exists()

    back_url = request.GET.get('next')
    if not back_url:
        from django.urls import reverse
        back_url = reverse('home')

    return render(request, 'product_detail.html', {
        'product': product,
        'can_save': can_save,
        'is_admin': is_admin,
        'is_store_owner': is_store_owner,
        'saved': saved,
        'back_url': back_url,
    })

def send_email_notification(to_email, product_name, store_name, sale_time):
    subject = f'แจ้งเตือนล่วงหน้า: สินค้า "{product_name}" จะเริ่มจำหน่ายเร็วๆ นี้!'
    message = f'''
    สวัสดีค่ะ,

    สินค้า "{product_name}" จากร้าน "{store_name}" จะเริ่มจำหน่ายในเวลา {sale_time.strftime('%H:%M')} น.

    กรุณาเตรียมตัวล่วงหน้า!

    ขอบคุณที่ใช้บริการของเรา 
    ทีมงาน YellowDeals
    '''
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [to_email])


@require_POST
def toggle_save_product(request, product_id):
    if request.method == 'POST':
        username = request.session.get('username')
        user = GeneralUser.objects.filter(customer_username=username).first()

        if not user:
            messages.error(request, "คุณต้องเข้าสู่ระบบในฐานะผู้ใช้ทั่วไปเพื่อบันทึกสินค้า")
            return redirect('product_detail', product_id=product_id)

        product = get_object_or_404(Product, product_id=product_id)

        # ตรวจสอบว่ามีการบันทึกสินค้าไว้หรือยัง
        saved_product = SavedProduct.objects.filter(customer=user, product=product).first()

        if saved_product:
            # ยกเลิกการบันทึก
            saved_product.delete()
            messages.success(request, "ยกเลิกการบันทึกสินค้าเรียบร้อยแล้ว")
        else:
            # ตรวจสอบเวลาเพื่อส่งอีเมลถ้าใกล้เปิดขาย
            now = timezone.now()
            if product.start_time and product.start_time - timedelta(hours=1) <= now <= product.start_time:
                send_email_notification(user.email, product.product_name, product.store_owner.store_name, product.start_time)

            # สร้างบันทึกสินค้าใหม่
            SavedProduct.objects.create(
                customer=user,
                product=product,
                saved_date=timezone.now()
            )
            messages.success(request, "บันทึกสินค้าสำเร็จ! คุณจะได้รับอีเมลแจ้งเตือนล่วงหน้า 1 ชั่วโมงก่อนขาย")

    return redirect('product_detail', product_id=product_id)

#-------เข้าสู่ระบบ--------
def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            admin = Admin.objects.filter(admin_username=username).first()
            customer = GeneralUser.objects.filter(customer_username=username).first()
            store_owner = StoreOwner.objects.filter(store_username=username).first()

            user_found = admin or customer or store_owner

            if not user_found:
                messages.error(request, 'ไม่พบชื่อผู้ใช้นี้ในระบบ กรุณาตรวจสอบอีกครั้ง')
            else:
                if admin:
                    if check_password(password, admin.admin_pass):
                        admin.last_login = now()  # อัพเดท last_login
                        admin.save()

                        request.session['username'] = admin.admin_username
                        messages.success(request, f'ยินดีต้อนรับแอดมิน {admin.admin_username}!')
                        return redirect('admin_dashboard')
                    else:
                        messages.error(request, 'รหัสผ่านไม่ถูกต้อง กรุณาลองใหม่')

                elif customer:
                    if check_password(password, customer.customer_pass):
                        customer.last_login = now()  # อัพเดท last_login
                        customer.save()

                        request.session['username'] = customer.customer_username
                        messages.success(request, f'สวัสดีคุณ {customer.customer_username}!')
                        return redirect('home')
                    else:
                        messages.error(request, 'รหัสผ่านไม่ถูกต้อง กรุณาลองใหม่')

                elif store_owner:
                    if store_owner.status == 'pending':
                        messages.warning(request, 'ขออภัย ขณะนี้ทางแอดมินกำลังตรวจสอบ กรุณาลองใหม่ภายใน 1-2 วัน')
                    elif store_owner.status == 'banned':
                        ban_reason = store_owner.note if store_owner.note else 'กรุณาติดต่อแอดมินสำหรับข้อมูลเพิ่มเติม'
                        messages.error(request, f'ร้านของคุณถูกระงับ: {ban_reason}')
                    elif check_password(password, store_owner.store_pass):
                        store_owner.last_login = now()  # อัพเดท last_login
                        store_owner.save()

                        request.session['username'] = store_owner.store_username
                        messages.success(request, f'เข้าสู่ระบบสำเร็จ! ยินดีต้อนรับเจ้าของร้าน {store_owner.store_username}!')
                        return redirect('store_dashboard')
                    else:
                        messages.error(request, 'รหัสผ่านไม่ถูกต้อง กรุณาลองใหม่')
        else:
            messages.error(request, 'กรุณากรอกข้อมูลให้ครบถ้วน')
    else:
        form = CustomAuthenticationForm()

    return render(request, 'login.html', {'form': form})

#-----ลงทะเบียนแอดมิน------
def admin_register_view(request):
    if not request.session.get('admin_code_verified'):
        return redirect('admin_secret')

    initial_data = {}
    if request.method == 'POST':
        form = AdminRegisterForm(request.POST)
        username = request.POST.get('admin_username')
        email = request.POST.get('admin_email')
        raw_password = request.POST.get('password')

        # เก็บค่าข้อมูลฟอร์มที่กรอก
        initial_data = {
            'admin_username': username,
            'admin_email': email
        }

        # เช็คว่ามีอีเมลซ้ำหรือไม่
        if Admin.objects.filter(admin_email=email).exists():
            request.session['form_error'] = {
                'type': 'error',
                'title': 'อีเมลถูกใช้แล้ว',
                'text': 'อีเมลนี้ถูกใช้งานแล้ว กรุณาใช้เมลใหม่'
            }
            request.session['form_data'] = initial_data
            return redirect('admin_register')

        # เช็ครหัสผ่าน
        if len(raw_password) < 8 or not re.search(r'[A-Za-z]', raw_password) or not re.search(r'[0-9]', raw_password):
            request.session['form_error'] = {
                'type': 'error',
                'title': 'รหัสผ่านไม่ปลอดภัย',
                'text': 'รหัสผ่านต้องมีอย่างน้อย 8 ตัว ประกอบด้วยตัวอักษรและตัวเลข'
            }
            request.session['form_data'] = initial_data
            return redirect('admin_register')

        # ทำการสร้างบัญชีแอดมินใหม่
        hashed_password = make_password(raw_password)
        new_admin = Admin.objects.create(
            admin_id='A' + str(Admin.objects.count() + 1).zfill(5),
            admin_username=username,
            admin_pass=hashed_password,
            admin_email=email
        )
        new_admin.save()

        # ลบข้อมูลที่เกี่ยวข้องกับรหัสที่ตรวจสอบแล้ว
        del request.session['admin_code_verified']

        request.session['form_error'] = {
            'type': 'success',
            'title': 'ลงทะเบียนสำเร็จ',
            'text': 'ลงทะเบียนแอดมินสำเร็จ! กรุณาเข้าสู่ระบบ'
        }
        return redirect('login')

    else:
        form = AdminRegisterForm(initial=request.session.pop('form_data', {}))

    return render(request, 'admin_register.html', {'form': form})

#---หน้ายืนยันรหัสแอดมิน ---
SECRET_CODE = 'YellwDeals001'  # รหัสลับสำหรับเข้าหน้าสมัครแอดมิน

def admin_secret_view(request):
    attempt_count = request.session.get('attempt_count', 0)
    locked_time = request.session.get('locked_time')

    if locked_time:
        locked_until = datetime.fromisoformat(locked_time) + timedelta(minutes=1)
        now = datetime.now()

        if now < locked_until:
            remaining = int((locked_until - now).total_seconds())
            messages.error(request, f"คุณกรอกผิดครบ 3 ครั้ง กรุณารอ {remaining} วินาที ก่อนลองใหม่")
            return render(request, 'admin_code.html')
        else:
            # ครบเวลาแล้ว reset
            request.session['attempt_count'] = 0
            request.session.pop('locked_time', None)

    if request.method == 'POST':
        code = request.POST.get('secret_code', '').strip()

        if code == SECRET_CODE:
            request.session['admin_code_verified'] = True
            request.session['attempt_count'] = 0
            request.session.pop('locked_time', None)  # เคลียร์เวลา
            return redirect('admin_register')
        else:
            attempt_count += 1
            request.session['attempt_count'] = attempt_count

            if attempt_count >= 3:
                request.session['locked_time'] = datetime.now().isoformat()
                messages.error(request, "คุณกรอกรหัสผิดครบ 3 ครั้ง ระบบจะล็อกไว้ 1 นาที!")
            else:
                remaining_attempts = 3 - attempt_count
                messages.error(request, f"รหัสพิเศษไม่ถูกต้อง! คุณเหลือโอกาสอีก {remaining_attempts} ครั้ง")

            return redirect('admin_secret')

    return render(request, 'admin_code.html')

#------ลงทะเบียนลูกค้า-----
def register_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        raw_password = request.POST['password']
        email = request.POST['email']

        errors = {}

        # ตรวจสอบว่าอีเมลซ้ำไหม
        if GeneralUser.objects.filter(customer_email=email).exists():
            errors['email'] = "อีเมลนี้ถูกใช้งานแล้ว กรุณาใช้เมลใหม่"

        # ตรวจสอบความยาวรหัสผ่าน
        if len(raw_password) < 8:
            errors['password'] = "รหัสผ่านต้องมีความยาวอย่างน้อย 8 ตัวอักษร"

        # ตรวจสอบความปลอดภัยของรหัสผ่าน
        elif not re.search(r'[A-Za-z]', raw_password) or not re.search(r'[0-9]', raw_password):
            errors['password'] = "รหัสผ่านต้องประกอบด้วยตัวอักษรและตัวเลข"

        if errors:
            for field, message_text in errors.items():
                messages.error(request, message_text)
            return render(request, 'register.html', {
                'username': username,
                'email': email
            })

        hashed_password = make_password(raw_password)
        new_user = GeneralUser.objects.create(
            customer_id='C' + str(GeneralUser.objects.count()+1).zfill(5),
            customer_username=username,
            customer_pass=hashed_password,
            customer_email=email
        )
        new_user.save()
        messages.success(request, "ลงทะเบียนสำเร็จ! กรุณาเข้าสู่ระบบ")
        return redirect('login')

    return render(request, 'register.html')

#-----ลงทะเบียนร้านค้า------
def register_shop_view(request):
    if request.method == 'POST':
        form = StoreOwnerRegistrationForm(request.POST)
        if form.is_valid():
            store_username = form.cleaned_data['store_username']
            raw_password = form.cleaned_data['store_pass']
            store_name = form.cleaned_data['store_name']
            store_email = form.cleaned_data['store_email']
            store_tel = form.cleaned_data['store_tel']
            store_address = form.cleaned_data['store_address']
            store_branch = form.cleaned_data['store_branch']
            typeStore = form.cleaned_data['typeStore']

            # รวม address + สาขา เพื่อเพิ่มความแม่นยำ
            full_address = f"{store_address} {store_branch}"
            lat, lng = get_lat_lng_from_address(full_address)

            hashed_password = make_password(raw_password)
            store_owner_id = 'S' + str(StoreOwner.objects.count() + 1).zfill(5)

            new_store = StoreOwner.objects.create(
                store_owner_id=store_owner_id,
                store_username=store_username,
                store_pass=hashed_password,
                store_name=store_name,
                store_email=store_email,
                store_tel=store_tel,
                store_address=store_address,
                store_branch=store_branch,
                typeStore=typeStore,
                status='pending',  # รออนุมัติ

                # เพิ่มพิกัด lat/lng
                latitude=lat,
                longitude=lng
            )
            new_store.save()

            messages.success(request, "ลงทะเบียนร้านค้าสำเร็จ! กรุณารอการอนุมัติจากแอดมิน ภายใน 1-2 วันหลังจากลงทะเบียน")
            return redirect('login')
    else:
        form = StoreOwnerRegistrationForm()

    return render(request, 'registershop.html', {'form': form})
#---รีเซ็ท PassWord----
def send_password_reset_email(request, user, email_field, password_field):
    # ตรวจสอบว่า `email_field` ที่ส่งมาให้มีฟิลด์อีเมลที่ถูกต้อง
    email = getattr(user, email_field, None)
    
    if not email:
        raise ValueError(f"{user.__class__.__name__} does not have a valid email field")

    # ตรวจสอบฟิลด์รหัสผ่านที่ถูกต้อง
    password = getattr(user, password_field, None)
    if not password:
        raise ValueError(f"{user.__class__.__name__} does not have the specified password field")

    # เช็คชื่อผู้ใช้จากแต่ละ Model
    if hasattr(user, 'admin_username'):
        username = user.admin_username
    elif hasattr(user, 'customer_username'):
        username = user.customer_username
    elif hasattr(user, 'store_username'):
        username = user.store_username
    else:
        username = "ผู้ใช้"

    # สร้าง token สำหรับการรีเซ็ตรหัสผ่าน
    subject = "Password Reset Request"
    email_template_name = "password_reset_email.html"
    c = {
        "email": email,
        "username": username,  # เพิ่มชื่อนี้เข้า template
        "domain": get_current_site(request).domain,
        "site_name": "Yellow Deals",
        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
        "token": default_token_generator.make_token(user),
        "protocol": "http",
    }

    email_message = render_to_string(email_template_name, c)
    try:
        send_mail(subject, email_message, "yellowdeals265@gmail.com", [email])
        print("ส่งอีเมลไปที่:", email)  # << เพิ่มบรรทัดนี้
    except BadHeaderError:
        print("มีปัญหา header ผิดพลาด!")


def password_reset_request(request):
    if request.method == "POST":
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']

            # ตรวจสอบในตาราง Admin
            associated_admin = Admin.objects.filter(admin_email=email)
            # ตรวจสอบในตาราง GeneralUser
            associated_customer = GeneralUser.objects.filter(customer_email=email)
            # ตรวจสอบในตาราง StoreOwner
            associated_store = StoreOwner.objects.filter(store_email=email)

            # ถ้าอีเมลตรงกับผู้ใช้ในตารางไหนก็ตาม ส่งอีเมลรีเซ็ตรหัสผ่าน
            if associated_admin.exists():
                for user in associated_admin:
                    send_password_reset_email(request, user, 'admin_email', 'admin_pass')
            if associated_customer.exists():
                for user in associated_customer:
                    send_password_reset_email(request, user, 'customer_email', 'customer_pass')
            if associated_store.exists():
                for user in associated_store:
                    send_password_reset_email(request, user, 'store_email', 'store_pass')

            # ส่งข้อความแจ้งเตือนหลังจากส่งอีเมลแล้ว
            if associated_admin.exists() or associated_customer.exists() or associated_store.exists():
                messages.success(request, "อีเมลได้รับการส่งไปพร้อมคำแนะนำสำหรับการรีเซ็ตรหัสผ่านของคุณแล้ว")
            else:
                messages.error(request, "อีเมลนี้ไม่ได้ลงทะเบียนในระบบ")
    else:
        form = PasswordResetForm()

    return render(request, "password_reset.html", {"form": form})
#-----ตั้งรหัสผ่านใหม่ ---
def password_reset_confirm(request, uidb64, token):
    user = None

    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        
        user = GeneralUser.objects.filter(customer_id=uid).first()
        if not user:
            user = Admin.objects.filter(admin_id=uid).first()
        if not user:
            user = StoreOwner.objects.filter(store_id=uid).first()

        if not user:
            messages.error(request, "ไม่พบผู้ใช้")
            return redirect('password_reset')

        if not default_token_generator.check_token(user, token):
            messages.error(request, "ลิงก์รีเซ็ตรหัสผ่านไม่ถูกต้องหรือหมดอายุ")
            return redirect("password_reset")

        if request.method == "POST":
            form = CustomSetPasswordForm(request.POST)
            if form.is_valid():
                new_password = form.cleaned_data['new_password1']
                hashed_password = make_password(new_password)

                # อัพเดต password hash ลงตาราง
                if hasattr(user, 'admin_pass'):
                    user.admin_pass = hashed_password
                elif hasattr(user, 'customer_pass'):
                    user.customer_pass = hashed_password
                elif hasattr(user, 'store_pass'):
                    user.store_pass = hashed_password

                user.save()
                messages.success(request, "เปลี่ยนรหัสผ่านเรียบร้อยแล้ว! คุณสามารถเข้าสู่ระบบด้วยรหัสผ่านใหม่ได้แล้ว!")
                return redirect("login")
        else:
            form = CustomSetPasswordForm()

    except (TypeError, ValueError, OverflowError):
        messages.error(request, "ข้อมูลผิดพลาด")
        return redirect("password_reset")

    return render(request, "password_reset_confirm.html", {"form": form})

#-----ออกจากระบบ ---
def logout_view(request):
    logout(request)
    request.session.flush()

    messages.success(request, 'คุณได้ออกจากระบบเรียบร้อยแล้ว')
    return redirect('home')
#-----หน้าหลักแอดมิน------
def admin_dashboard(request):
    username = request.session.get('username')

    if not username:
        return redirect('login')  # ถ้าไม่มี session กลับไปล็อกอิน

    admin = Admin.objects.filter(admin_username=username).first()
    
    if not admin:
        return redirect('login')  # กันไว้กรณีชื่อใน session ไม่ตรง database

    return render(request, 'admin_dashboard.html', {'admin': admin})

#-----หน้าจัดการคำขอ (แอดมิน)------
def admin_ManageRequests(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    date = request.GET.get('date')
    sort_order = request.GET.get('sort', 'all')

    store_list = StoreOwner.objects.filter(status='pending')

    if date:
        store_list = store_list.filter(register_date=date)

    if sort_order == 'oldest':
        store_list = store_list.order_by('register_date')
    elif sort_order == 'latest':
        store_list = store_list.order_by('-register_date')
    else:
        store_list = store_list.order_by('store_name')

    paginator = Paginator(store_list, 3)  # ✅ แสดงหน้าละ 3 รายการ
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'admin_ManageRequests.html', {
        'admin': admin,
        'store_list': page_obj,  # ✅ ใช้ page_obj แทน
        'selected_date': date,
        'sort_order': sort_order,
    })
# --- เพิ่มฟังก์ชันแปลงวันที่ไทย ---
def format_thai_date(date):
    if not date:
        return "-"
    
    thai_months = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    day = date.day
    month = thai_months[date.month]
    year = date.year + 543  # แปลง ค.ศ. ➔ พ.ศ.
    return f"{day} {month} {year}"
# --- แสดงรายละเอียดร้านค้า---
def admin_store_detail(request, store_id):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    store = get_object_or_404(StoreOwner, store_owner_id=store_id)

    # --- เพิ่ม field วันที่ไทยเข้าไปใน store ---
    store.register_date_thai = format_thai_date(store.register_date)

    return render(request, 'admin_store_detail.html', {
        'store': store,
        'admin': admin,  # ส่ง admin กลับไปด้วย
    })
#------แปลงที่อยู่อัตโนมัติ---
def clean_address(address):
    address = address.replace("รหัสสาขา 01800", "")
    address = address.replace("รหัสสาขา", "")
    return address.strip()

# ฟังก์ชันสำหรับหาพิกัดจากที่อยู่
def get_lat_lng_from_address(address):
    address = clean_address(address)  # ทำความสะอาดที่อยู่
    encoded_address = urllib.parse.quote(address)
    
    # ใช้ Nominatim API เป็นหลัก
    lat, lng = get_lat_lng_from_nominatim(encoded_address)
    if lat is not None and lng is not None:
        return lat, lng
    
    # ถ้า Nominatim หาไม่ได้
    return None, None

# ฟังก์ชันสำหรับหาพิกัดจาก Nominatim API
def get_lat_lng_from_nominatim(encoded_address):
    url = f'https://nominatim.openstreetmap.org/search?q={encoded_address}&format=json'
    response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
    result = response.json()
    
    if result:
        lat = float(result[0]['lat'])
        lng = float(result[0]['lon'])
        return lat, lng
    return None, None

# ฟังก์ชันอนุมัติร้านค้า
def approve_store(request, store_id):
    username = request.session.get('username')
    
    if not Admin.objects.filter(admin_username=username).exists():
        return redirect('login')

    try:
        store = StoreOwner.objects.get(store_owner_id=store_id)

        # ตรวจสอบอีเมลร้าน
        try:
            validate_email(store.store_email)
        except ValidationError:
            messages.error(request, 'อีเมลของร้านไม่ถูกต้อง ไม่สามารถส่งเมลได้')
            return redirect('admin_ManageRequests')

        store.status = 'active'

        # หาพิกัด
        lat, lng = get_lat_lng_from_address(store.store_address)
        if lat is not None and lng is not None:
            store.latitude = lat
            store.longitude = lng
        else:
            messages.warning(request, 'ไม่สามารถระบุพิกัดจากที่อยู่ได้')

        store.save()

        # ส่งอีเมลแจ้งเตือน
        email_body = render_to_string('approve_store_email.html', {
            'store_name': store.store_name
        })
        subject = 'ร้านค้าของคุณได้รับการอนุมัติแล้ว'
        try:
            send_mail(subject, '', settings.EMAIL_HOST_USER, [store.store_email], html_message=email_body, fail_silently=False)
        except BadHeaderError:
            messages.error(request, 'เกิดข้อผิดพลาดในการส่งอีเมล')

        messages.success(request, f'อนุมัติร้าน "{store.store_name}" แล้วเรียบร้อย')
    except StoreOwner.DoesNotExist:
        messages.error(request, 'ไม่พบร้านที่ต้องการอนุมัติ')

    return redirect('admin_ManageRequests')
# ปฏิเสธร้าน
def reject_store(request, store_id):
    username = request.session.get('username')
    if not Admin.objects.filter(admin_username=username).exists():
        return redirect('login')

    try:
        store = StoreOwner.objects.get(store_owner_id=store_id)

        try:
            validate_email(store.store_email)
        except ValidationError:
            messages.error(request, 'อีเมลของร้านไม่ถูกต้อง ไม่สามารถส่งเมลได้')
            return redirect('admin_ManageRequests')

        # โหลดเนื้อหาอีเมลจากเทมเพลต
        email_body = render_to_string('reject_store_email.html', {
            'store_name': store.store_name
        })

        # ส่งอีเมล
        subject = 'ร้านค้าของคุณไม่ได้รับการอนุมัติ'
        try:
            send_mail(subject, '', settings.EMAIL_HOST_USER, [store.store_email], html_message=email_body, fail_silently=False)
        except BadHeaderError:
            messages.error(request, 'เกิดข้อผิดพลาดในการส่งอีเมล')

        store.delete()
        messages.success(request, f'ปฏิเสธร้าน "{store.store_name}" แล้วเรียบร้อย')
    except StoreOwner.DoesNotExist:
        messages.error(request, 'ไม่พบร้านที่ต้องการปฏิเสธ')

    return redirect('admin_ManageRequests')

#-----หน้าตรวจสอบคำร้องเรียน (แอดมิน)------
def admin_complaints(request):
    username = request.session.get('username')

    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()

    if not admin:
        return redirect('login')

    # ดึงประเภทคำร้องเรียนทั้งหมดจาก TypeComplaint
    type_complaints = TypeComplaint.objects.all()

    # ดึงคำร้องเรียนทั้งหมด
    complaint_list = Complaint.objects.select_related('customer', 'store_owner', 'type_complaint').order_by('-complaint_date')

    # รับค่าจากฟอร์ม
    date = request.GET.get('date')
    type_complaint_id = request.GET.get('type_complaint')  # <- เปลี่ยนชื่อให้ตรงกับ name ในฟอร์ม
    role = request.GET.get('role')

    # กรองตามวันที่
    if date:
        complaint_list = complaint_list.filter(complaint_date__date=date)

    # กรองตามประเภทคำร้องเรียน
    if type_complaint_id:
        complaint_list = complaint_list.filter(type_complaint__typeComplaint_id=type_complaint_id)

    # กรองตามบทบาท
    if role == 'user':
        complaint_list = complaint_list.filter(customer__isnull=False)
    elif role == 'store':
        complaint_list = complaint_list.filter(store_owner__isnull=False)

    total_complaints = complaint_list.count()

    return render(request, 'admin_complaints.html', {
        'admin': admin,
        'type_complaints': type_complaints,
        'complaint_list': complaint_list,
        'total_complaints': total_complaints,
    })
#----- หน้ารายละเอียดคำร้องเรียน -----
def complaint_detail(request, complaint_id):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    complaint = get_object_or_404(Complaint, pk=complaint_id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update':
            new_status = request.POST.get('status')
            if new_status:  # ตรวจสอบว่ามีการเลือกสถานะจริง ๆ
                complaint.complaint_status = new_status
                complaint.save()
                messages.success(request, "อัปเดตสถานะเรียบร้อยแล้ว")
            else:
                messages.warning(request, "กรุณาเลือกสถานะก่อนอัปเดต")

        elif action == 'delete':
            complaint.delete()
            messages.success(request, "ลบคำร้องเรียนเรียบร้อยแล้ว")
            return redirect('admin_complaints')  # เปลี่ยนเป็นหน้าแสดงรายการคำร้อง

    return render(request, 'complaint_detail.html', {
        'admin': admin,
        'complaint': complaint
    })

#-----หน้าตรวจสอบร้านค้า (แอดมิน)------
def admin_review_shop(request, store_owner_id):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    store = get_object_or_404(StoreOwner, store_owner_id=store_owner_id)

    return render(request, 'admin_review_shop.html', {
        'admin': admin,
        'store': store
    })

def admin_shops(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    # pop alert จาก session ถ้ามี
    alert = request.session.pop('alert', None)

    # รับค่าจาก GET
    shop_name = request.GET.get('shop_name', '').strip()
    type_store_id = request.GET.get('type_store', '')
    status_filter = request.GET.get('status', 'all')  # ค่าเริ่มต้นเป็น 'all'

    # เริ่มจากร้านที่มีสถานะ active หรือ banned เท่านั้น
    store_list = StoreOwner.objects.filter(status__in=['active', 'banned'])

    # ถ้าเลือก filter สถานะ
    if status_filter != 'all':
        store_list = store_list.filter(status=status_filter)

    # กรองชื่อร้าน
    if shop_name:
        store_list = store_list.filter(store_name__icontains=shop_name)

    # กรองตามประเภทร้าน
    if type_store_id:
        store_list = store_list.filter(typeStore_id=type_store_id)

    # เรียงตามลำดับล่าสุด
    store_list = store_list.order_by('-store_owner_id')

    # ดึง typeStore ทั้งหมดสำหรับ dropdown
    type_stores = TypeStore.objects.all()

    return render(request, 'admin_shops.html', {
        'admin': admin,
        'store_list': store_list,
        'shop_name': shop_name,
        'type_stores': type_stores,
        'selected_status': status_filter,
        'selected_type_store': type_store_id,
        'alert': alert,  # ส่ง alert เข้า template
    })
#-----หน้าแบนหรือไม่แบน (แอดมิน)------
@require_POST
def ban_shop(request, store_id):
    store = get_object_or_404(StoreOwner, pk=store_id)
    note = request.POST.get('note', '').strip()

    # ตรวจสอบหมายเหตุ
    if not note:
        request.session['alert'] = {
            'type': 'error',
            'title': 'เกิดข้อผิดพลาด',
            'text': 'กรุณาระบุหมายเหตุในการแบนร้านค้า'
        }
        return redirect('admin_shops')

    # อัปเดตสถานะร้านเป็น banned
    store.status = 'banned'
    store.note = note
    store.save()

    # แจ้งเตือนการแบนร้าน
    request.session['alert'] = {
        'type': 'success',
        'title': 'สำเร็จ',
        'text': f"ร้าน {store.store_name} ถูกแบนแล้ว"
    }

    return redirect('admin_shops')

@require_POST
def unban_shop(request, store_id):
    store = get_object_or_404(StoreOwner, pk=store_id)
    
    # ปลดแบนร้าน
    store.status = 'active'
    store.note = ''
    store.save()

    # แจ้งเตือนการปลดแบนร้าน
    request.session['alert'] = {
        'type': 'success',
        'title': 'ปลดแบนร้านสำเร็จ',
        'text': f"ร้าน {store.store_name} ถูกปลดแบนแล้ว"
    }

    return redirect('admin_shops')

#-----หน้าแก้ไขรายละเอียด (แอดมิน)------
def admin_edit(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    # ดึงข้อมูล
    type_complaints = TypeComplaint.objects.all()
    type_products = TypeProduct.objects.all()
    type_stores = TypeStore.objects.all()
    categories = Categories.objects.all()

    # ดึง filter parameter สำหรับพิกัดร้าน
    filter_option = request.GET.get('filter', 'all')

    store_owners = StoreOwner.objects.all()

    # กรองข้อมูล store_owners ตาม filter_option
    if filter_option == 'no_location':
        # ยังไม่มีพิกัด (latitude หรือ longitude เป็น NULL)
        store_owners = store_owners.filter(
            Q(latitude__isnull=True) | Q(longitude__isnull=True)
        )
    elif filter_option == 'has_location':
        # มีพิกัดครบ (latitude และ longitude ไม่เป็น NULL)
        store_owners = store_owners.filter(
            latitude__isnull=False,
            longitude__isnull=False
        )

    # Pagination
    paginator_complaints = Paginator(type_complaints, 3)
    paginator_products = Paginator(type_products, 3)
    paginator_stores = Paginator(type_stores, 3)
    paginator_categories = Paginator(categories, 3)
    paginator_store_owners = Paginator(store_owners, 3)

    # Current page
    page_complaints = request.GET.get('page_complaints')
    page_products = request.GET.get('page_products')
    page_stores = request.GET.get('page_stores')
    page_categories = request.GET.get('page_categories')
    page_store_owners = request.GET.get('page_store_owners')

    # Get paginated data
    type_complaints_page = paginator_complaints.get_page(page_complaints)
    type_products_page = paginator_products.get_page(page_products)
    type_stores_page = paginator_stores.get_page(page_stores)
    categories_page = paginator_categories.get_page(page_categories)
    store_owners_page = paginator_store_owners.get_page(page_store_owners)

    # Success message
    success_message = request.GET.get('success_message')

    return render(request, 'admin_edit.html', {
        'admin': admin,
        'type_complaints': type_complaints_page,
        'type_products': type_products_page,
        'type_stores': type_stores_page,
        'categories': categories_page,
        'store_owners': store_owners_page,
        'success_message': success_message,
        'filter_option': filter_option,
    })

#แก้พิกัดร้าน 
def edit_store_location(request, store_id):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')
    store = get_object_or_404(StoreOwner, pk=store_id)

    if request.method == 'POST':
        store.latitude = request.POST.get('latitude')
        store.longitude = request.POST.get('longitude')
        store.save()

        # ✅ แจ้งเตือน
        messages.success(request, 'แก้ไขพิกัดเรียบร้อยแล้ว')

        # ✅ redirect กลับไปยัง next
        next_url = request.POST.get('next') or 'admin_edit'
        return redirect(next_url)

    # GET: เตรียม next สำหรับใส่ใน template
    next_url = request.GET.get('next', 'admin_edit')
    return render(request, 'edit_store_location.html', {
        'admin': admin,
        'store': store,
        'next': next_url,
    })
# เพิ่มข้อมูลประเภทคำร้อง
def add_complaint(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = TypeComplaintForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('typeComplaint_name')
            if TypeComplaint.objects.filter(typeComplaint_name=name).exists():
                messages.error(request, "ชื่อประเภทคำร้องนี้มีอยู่แล้ว")
            else:
                form.save()
                messages.success(request, "เพิ่มข้อมูลประเภทคำร้องสำเร็จ")
                return redirect(next_url)
    else:
        form = TypeComplaintForm()
    return render(request, 'add_complaint.html', {'form': form, 'previous_url': next_url})

# เพิ่มข้อมูลประเภทสินค้า
def add_typeproduct(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = TypeProductForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('productType_name')
            if TypeProduct.objects.filter(productType_name=name).exists():
                messages.error(request, "ชื่อประเภทสินค้านี้มีอยู่แล้ว")
            else:
                form.save()
                messages.success(request, "เพิ่มข้อมูลประเภทสินค้าสำเร็จ")
                return redirect(next_url)
    else:
        form = TypeProductForm()
    return render(request, 'add_typeproduct.html', {'form': form, 'previous_url': next_url})

# เพิ่มข้อมูลประเภทของร้าน
def generate_type_store_id():
    last_store = TypeStore.objects.order_by('-typeStore_id').first()
    if last_store:
        last_id = int(last_store.typeStore_id[1:])
        new_id = last_id + 1
    else:
        new_id = 1
    return f'B{new_id:05d}'

def add_store(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = TypeStoreForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('typeStore_name')
            if TypeStore.objects.filter(typeStore_name=name).exists():
                messages.error(request, "ชื่อประเภทของร้านนี้มีอยู่แล้ว")
            else:
                new_id = generate_type_store_id()
                TypeStore.objects.create(
                    typeStore_id=new_id,
                    typeStore_name=name
                )
                messages.success(request, "เพิ่มข้อมูลประเภทของร้านสำเร็จ")
                return redirect(next_url)
    else:
        form = TypeStoreForm()
    return render(request, 'add_store.html', {'form': form, 'previous_url': next_url})

# เพิ่มข้อมูลหมวดหมู่
def add_category(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = CategoriesForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('categories_name')
            if Categories.objects.filter(categories_name=name).exists():
                messages.error(request, "ชื่อหมวดหมู่นี้มีอยู่แล้ว")
            else:
                form.save()
                messages.success(request, "เพิ่มข้อมูลหมวดหมู่สำเร็จ")
                return redirect(next_url)
    else:
        form = CategoriesForm()
    return render(request, 'add_category.html', {'form': form, 'previous_url': next_url})

# ลบข้อมูล
def delete_complaint(request, pk):
    try:
        complaint = get_object_or_404(TypeComplaint, pk=pk)
        complaint.delete()
        messages.success(request, "ลบข้อมูลประเภทคำร้องสำเร็จ")
    except Exception as e:
        messages.error(request, f"เกิดข้อผิดพลาดในการลบข้อมูล: {str(e)}")
    return redirect(request.GET.get('next', 'admin_edit'))

def delete_typeproduct(request, pk):
    product = get_object_or_404(TypeProduct, pk=pk)
    product.delete()
    messages.success(request, "ลบข้อมูลประเภทสินค้าสำเร็จ")
    return redirect(request.GET.get('next', 'admin_edit'))

def delete_store(request, pk):
    store = get_object_or_404(TypeStore, pk=pk)
    store.delete()
    messages.success(request, "ลบข้อมูลประเภทของร้านสำเร็จ")
    return redirect(request.GET.get('next', 'admin_edit'))

def delete_category(request, pk):
    category = get_object_or_404(Categories, pk=pk)
    category.delete()
    messages.success(request, "ลบข้อมูลหมวดหมู่สำเร็จ")
    return redirect(request.GET.get('next', 'admin_edit'))

# แก้ไขข้อมูล
def edit_complaint(request, pk):
    complaint = get_object_or_404(TypeComplaint, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = TypeComplaintForm(request.POST, instance=complaint)
        if form.is_valid():
            name = form.cleaned_data.get('typeComplaint_name')
            if TypeComplaint.objects.exclude(pk=pk).filter(typeComplaint_name=name).exists():
                messages.error(request, "ชื่อประเภทคำร้องนี้มีอยู่แล้ว")
            else:
                form.save()
                messages.success(request, "แก้ไขข้อมูลประเภทคำร้องสำเร็จ")
                return redirect(next_url)
    else:
        form = TypeComplaintForm(instance=complaint)
    return render(request, 'edit_complaint.html', {'form': form, 'previous_url': next_url})

def edit_typeproduct(request, pk):
    product = get_object_or_404(TypeProduct, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = TypeProductForm(request.POST, instance=product)
        if form.is_valid():
            name = form.cleaned_data.get('productType_name')
            if TypeProduct.objects.exclude(pk=pk).filter(productType_name=name).exists():
                messages.error(request, "ชื่อประเภทสินค้านี้มีอยู่แล้ว")
            else:
                form.save()
                messages.success(request, "แก้ไขข้อมูลประเภทสินค้าสำเร็จ")
                return redirect(next_url)
    else:
        form = TypeProductForm(instance=product)
    return render(request, 'edit_typeproduct.html', {'form': form, 'previous_url': next_url})

def edit_store(request, pk):
    store = get_object_or_404(TypeStore, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = TypeStoreForm(request.POST, instance=store)
        if form.is_valid():
            name = form.cleaned_data.get('typeStore_name')
            if TypeStore.objects.exclude(pk=pk).filter(typeStore_name=name).exists():
                messages.error(request, "ชื่อประเภทของร้านนี้มีอยู่แล้ว")
            else:
                form.save()
                messages.success(request, "แก้ไขข้อมูลประเภทของร้านสำเร็จ")
                return redirect(next_url)
    else:
        form = TypeStoreForm(instance=store)
    return render(request, 'edit_store.html', {'form': form, 'previous_url': next_url})

def edit_category(request, pk):
    category = get_object_or_404(Categories, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or 'admin_edit'
    if request.method == 'POST':
        form = CategoriesForm(request.POST, instance=category)
        if form.is_valid():
            name = form.cleaned_data.get('categories_name')
            if Categories.objects.exclude(pk=pk).filter(categories_name=name).exists():
                messages.error(request, "ชื่อหมวดหมู่นี้มีอยู่แล้ว")
            else:
                form.save()
                messages.success(request, "แก้ไขข้อมูลหมวดหมู่สำเร็จ")
                return redirect(next_url)
    else:
        form = CategoriesForm(instance=category)
    return render(request, 'edit_category.html', {'form': form, 'previous_url': next_url})
#-----หน้าข้อมูลของฉัน (แอดมิน)------
def admin_profile_view(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    success = request.GET.get('success') == '1'

    return render(request, 'admin_profile_view.html', {
        'admin': admin,
        'success': success,
    })

def admin_profile(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    admin = Admin.objects.filter(admin_username=username).first()
    if not admin:
        return redirect('login')

    next_url = request.GET.get('next') or request.POST.get('next')

    if request.method == 'POST':
        new_username = request.POST.get('admin_username')
        new_email = request.POST.get('admin_email')
        new_tel = request.POST.get('admin_tel')

        admin.admin_username = new_username
        admin.admin_email = new_email
        admin.admin_tel = new_tel
        admin.save()

        # อัปเดต session ด้วย username ใหม่
        request.session['username'] = new_username

        # แก้ตรงนี้ ➔ เติม success=1 ใน next_url ถ้ายังไม่มี
        if next_url:
            if 'success=1' not in next_url:
                if '?' in next_url:
                    next_url += '&success=1'
                else:
                    next_url += '?success=1'
            return redirect(next_url)
        else:
            return redirect(reverse('admin_profile_view') + '?success=1')

    return render(request, 'admin_profile.html', {
        'admin': admin,
        'next_url': next_url or '',
    })
    
#-----หน้าหลักร้านค้า------
def store_dashboard(request):
    username = request.session.get('username')

    if not username:
        return redirect('login')

    store_owner = StoreOwner.objects.filter(store_username=username).first()

    if not store_owner:
        return redirect('login')

    context = {
        'StoreOwner': store_owner, 
    }
    return render(request, 'store_dashboard.html', context)

#-----หน้าจัดการสินค้า (ร้านค้า)------
def store_manage_product(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    store_owner = StoreOwner.objects.filter(store_username=username).first()
    if not store_owner:
        return redirect('login')
    
    now = timezone.now()

    all_qs       = Product.objects.filter(store_owner=store_owner).order_by('-product_date')
    active_qs    = all_qs.filter(start_time__lte=now, end_time__gte=now, is_hidden=False)
    inactive_qs  = all_qs.filter(end_time__lt=now, is_hidden=False)
    hidden_qs    = all_qs.filter(is_hidden=True)

    paginator_all      = Paginator(all_qs, 4)
    paginator_active   = Paginator(active_qs, 4)
    paginator_inactive = Paginator(inactive_qs, 4)
    paginator_hidden   = Paginator(hidden_qs, 4)

    tab = request.GET.get('tab', 'all')
    page = request.GET.get('page', 1)

    if tab == 'all':
        all_products_page = paginator_all.get_page(page)
        active_products_page = paginator_active.get_page(1)
        inactive_products_page = paginator_inactive.get_page(1)
        hidden_products_page = paginator_hidden.get_page(1)
    elif tab == 'active':
        all_products_page = paginator_all.get_page(1)
        active_products_page = paginator_active.get_page(page)
        inactive_products_page = paginator_inactive.get_page(1)
        hidden_products_page = paginator_hidden.get_page(1)
    elif tab == 'inactive':
        all_products_page = paginator_all.get_page(1)
        active_products_page = paginator_active.get_page(1)
        inactive_products_page = paginator_inactive.get_page(page)
        hidden_products_page = paginator_hidden.get_page(1)
    elif tab == 'hidden':
        all_products_page = paginator_all.get_page(1)
        active_products_page = paginator_active.get_page(1)
        inactive_products_page = paginator_inactive.get_page(1)
        hidden_products_page = paginator_hidden.get_page(page)
    else:
        # ค่าเริ่มต้น
        all_products_page = paginator_all.get_page(1)
        active_products_page = paginator_active.get_page(1)
        inactive_products_page = paginator_inactive.get_page(1)
        hidden_products_page = paginator_hidden.get_page(1)

    context = {
        'StoreOwner': store_owner,
        'now': now,
        'all_products':      all_products_page,
        'active_products':   active_products_page,
        'inactive_products': inactive_products_page,
        'hidden_products':   hidden_products_page,
    }

    return render(request, 'store_manage_product.html', context)



#-----หน้าเพิ่ม ลบ แก้ไข สินค้า (ร้านค้า)------
def add_product(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'product_list'
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            store_owner = StoreOwner.objects.get(store_username=request.session.get('username'))
            product.store_owner = store_owner
            product.save()
            messages.success(request, "เพิ่มข้อมูลสินค้าสำเร็จ")
            return redirect(next_url)
    else:
        form = ProductForm()
    return render(request, 'add_product.html', {'form': form, 'previous_url': next_url})

# ลบสินค้า
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    try:
        product.delete()
        messages.success(request, "ลบข้อมูลสินค้าสำเร็จ")
    except Exception as e:
        messages.error(request, f"เกิดข้อผิดพลาดในการลบข้อมูล: {str(e)}")
    return redirect(request.GET.get('next', 'product_list'))

# แก้ไขสินค้า
def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or 'product_list'
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, "แก้ไขข้อมูลสินค้าสำเร็จ")
            return redirect(next_url)
    else:
        form = ProductForm(instance=product)
    return render(request, 'edit_product.html', {
        'form': form,
        'previous_url': next_url,
        'product': product,   
    })


#---หน้าการซ่อนสินค้า (ร้านค้า)----
def manage_products(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    store_owner = StoreOwner.objects.filter(store_username=username).first()
    if not store_owner:
        return redirect('login')

    store_owner_id = store_owner.store_owner_id

    active_products = Product.objects.filter(
        store_owner=store_owner, 
        is_hidden=False, 
        end_time__gte=timezone.now()
    )
    inactive_products = Product.objects.filter(
        store_owner=store_owner, 
        end_time__lte=timezone.now(),
        is_hidden=False
    )
    hidden_products = Product.objects.filter(
        store_owner=store_owner, 
        is_hidden=True
    )

    context = {
        'StoreOwner': store_owner,
        'active_products': active_products,
        'inactive_products': inactive_products,
        'hidden_products': hidden_products,
        'all_products': Product.objects.filter(store_owner=store_owner), 
        'now': timezone.now(),  
    }
    return render(request, 'store_manage_product.html', context)


def toggle_hidden(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.is_hidden = not product.is_hidden
    product.save()
    return redirect(request.GET.get('next', 'store_manage_product'))

#---หน้าการรายงาน (ร้านค้า)----
def store_reports(request):
    import locale
    locale.setlocale(locale.LC_TIME, 'th_TH.UTF-8')
    username = request.session.get('username')
    if not username:
        return redirect('login')

    store_owner = StoreOwner.objects.filter(store_username=username).first()
    if not store_owner:
        return redirect('login')

    now = timezone.now()
    filter_type = request.GET.get('filter', 'all')
    selected_month = request.GET.get('month')

    try:
        selected_month_int = int(selected_month) if selected_month else None
    except ValueError:
        selected_month_int = None

    if filter_type in ['month', 'productChart']:
        if selected_month_int:
            products_queryset = Product.objects.filter(
                store_owner=store_owner,
                product_date__month=selected_month_int,
                product_date__year=now.year
            ).order_by('-product_date')
        else:
            products_queryset = Product.objects.filter(
                store_owner=store_owner
            ).order_by('-product_date')
    else:
        products_queryset = Product.objects.filter(store_owner=store_owner).order_by('-product_date')

    total_views = products_queryset.aggregate(Sum('view_count'))['view_count__sum'] or 0

    paginator = Paginator(products_queryset, 3)
    page_number = request.GET.get('page')
    page_number_int = int(page_number) if page_number and page_number.isdigit() else 1
    products = paginator.get_page(page_number)

    start_num = (page_number_int - 1) * paginator.per_page

    months_thai = [
        (1, "มกราคม"), (2, "กุมภาพันธ์"), (3, "มีนาคม"), (4, "เมษายน"),
        (5, "พฤษภาคม"), (6, "มิถุนายน"), (7, "กรกฎาคม"), (8, "สิงหาคม"),
        (9, "กันยายน"), (10, "ตุลาคม"), (11, "พฤศจิกายน"), (12, "ธันวาคม")
    ]

    context = {
        'StoreOwner': store_owner,
        'products': products,
        'now': now,
        'filter_type': filter_type,
        'months_thai': months_thai,
        'selected_month': selected_month,
        'total_views': total_views,
        'page_number': page_number_int,
        'per_page': paginator.per_page,
        'start_num': start_num,
    }
    return render(request, 'store_reports.html', context)


#---หน้ากราฟ/EXCAL(ร้านค้า)----
def export_product_report(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    store_owner = StoreOwner.objects.filter(store_username=username).first()
    if not store_owner:
        return redirect('login')

    selected_month = request.GET.get('month')
    now = timezone.now()

    try:
        selected_month_int = int(selected_month) if selected_month else None
    except ValueError:
        selected_month_int = None

    # กรองตามเดือน ถ้ามี
    if selected_month_int:
        products = Product.objects.filter(
            store_owner=store_owner,
            product_date__month=selected_month_int,
            product_date__year=now.year
        )
    else:
        products = Product.objects.filter(store_owner=store_owner)

    wb = Workbook()
    ws = wb.active
    ws.title = 'รายงานสินค้า'

    headers = ['ลำดับ', 'ชื่อสินค้า', 'จำนวนการเข้าชม', 'ลดราคาเมื่อ']
    ws.append(headers)

    column_widths = [10, 30, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for idx, product in enumerate(products, 1):
        if product.start_time:
            start_time = product.start_time.strftime('%d/%m/%Y')
        else:
            start_time = 'ยังไม่กำหนดวันลดราคา'

        ws.append([idx, product.product_name, product.view_count, start_time])

    filename = "product_report.xlsx"
    filename = filename.encode('utf-8').decode('utf-8')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response
#---หน้าแก้ไขรายละเอียดร้านค้า (ร้านค้า) ---
def store_about(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')

    store_owner = StoreOwner.objects.filter(store_username=username).first()
    if not store_owner:
        return redirect('login')

    context = {
        'store_owner': store_owner,
    }

    return render(request, 'store_about.html', context)

def store_profile_edit(request):
    username = request.session.get('username')
    store_owner = get_object_or_404(StoreOwner, store_username=username)

    if request.method == 'POST':
        # ดึงค่าชื่อร้านและอีเมลจาก hidden input เพราะ readonly input จะไม่ถูกส่งไป
        post_data = request.POST.copy()
        post_data['store_name'] = request.POST.get('store_name', store_owner.store_name)
        post_data['store_email'] = request.POST.get('store_email', store_owner.store_email)

        form = StoreProfileForm(post_data, request.FILES, instance=store_owner)

        if form.is_valid():
            form.save()
            messages.success(request, "บันทึกข้อมูลร้านเรียบร้อยแล้ว")
            return redirect('store_about')
        else:
            print("Form Errors:", form.errors)  # สำหรับ debug
    else:
        form = StoreProfileForm(instance=store_owner)

    return render(request, 'store_profile_edit.html', {
        'form': form,
        'StoreOwner': store_owner,
    })

def store_delete_account(request):
    username = request.session.get('username')
    store_owner = StoreOwner.objects.filter(store_username=username).first()

    next_url = request.GET.get('next') or 'login'

    if store_owner:
        store_owner.delete()
        request.session.flush()
        logout(request)

    return redirect(next_url)

